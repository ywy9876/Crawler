# -*- coding: utf-8 -*-
"""
Created on Fri Jan 20 23:38:34 2017

@author: ywy
"""

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import time
import openpyxl as op
import re


driver = webdriver.Chrome('path to chromedriver')
error_log = open('error_log.txt', 'w')

#只运行一次用作登录
driver.get("link to login page")
elem = driver.find_element_by_name("user")
elem.send_keys("your username")
elem = driver.find_element_by_name("pass")
elem.send_keys("your password")
elem.send_keys(Keys.RETURN)
driver.get('link of reset page')
time.sleep(5)

#第一次填写表单，随便填什么，日期来源选好就行        
driver.find_element_by_id('addRowLink').click() #添加行

time.sleep(1)
elem = driver.find_element_by_id('queryTermField_0')
if elem.is_displayed():
    elem.clear()
    elem.send_keys('billion')  #输入billion
else:
    time.sleep(1)
    elem.clear()
    elem.send_keys('billion')
    
time.sleep(1)
elem = driver.find_element_by_id('queryTermFieldRight_0')
if elem.is_displayed():
    elem.clear()
    elem.send_keys('million') #输入million
else:
    time.sleep(1)
    elem.clear()
    elem.send_keys('million')
    
select = Select(driver.find_element_by_id('fieldsSelect'))
select.select_by_value('ti')  #选择Document Title
select = Select(driver.find_element_by_id('fieldsSelect_0'))
select.select_by_value('ti')    

time.sleep(1)
elem = driver.find_element_by_id('queryTermField_1')
if elem.is_displayed():
    elem.clear()
    elem.send_keys('invest')
else:
    time.sleep(1)
    elem.clear()
    elem.send_keys('invest')
select = Select(driver.find_element_by_id('fieldsSelect_1'))
select.select_by_value('ti')    


select = Select(driver.find_element_by_id('select_multiDateRange'))
select.select_by_value('RANGE')  #选择出版日期为区间

select = Select(driver.find_element_by_id('month2'))  
select.select_by_value('JANUARY') #开始月份为一月
select = Select(driver.find_element_by_id('day2'))   
select.select_by_value('1') #开始日期为1号
driver.find_element_by_id('year2').send_keys('1996') #1996年起

select = Select(driver.find_element_by_id('month2_0')) 
select.select_by_value('DECEMBER') #结束月份为12月
select = Select(driver.find_element_by_id('day2_0'))
select.select_by_value('31') #结束日期为31日
driver.find_element_by_id('year2_0').send_keys('2001') #至2001

#等待SourceType选择框加载4秒
time.sleep(4)
driver.find_element_by_id('SourceType_Newspapers').click()
driver.find_element_by_id('SourceType_Wire_Feeds').click()

select = Select(driver.find_element_by_id('itemsPerPage'))
select.select_by_value('100')

time.sleep(1)
elem = driver.find_element_by_id('queryTermField')
if elem.is_displayed():
    elem.clear()
    elem.send_keys('IBM')  #输入IBM
else:
    time.sleep(1)
    elem.clear()
    elem.send_keys('IBM')
elem.send_keys(Keys.RETURN)
time.sleep(5)


wb = op.load_workbook('1.xlsx', use_iterators=True) #公司名excel
ws = wb.get_sheet_by_name('Sheet1')
wbres = op.load_workbook('result.xlsx')  #存结果的excel
if 'res' in wbres.get_sheet_names():
    wsres = wbres.get_sheet_by_name('res')
    wbres.remove_sheet(wsres)
wbres.create_sheet('res')
wsres = wbres.get_sheet_by_name('res')
pattern = re.compile('<li class="resultItem ltr".*?<a id="citationDocTitleLink".*?title="(.*?)" class.*?<span class="titleAuthorETC small".*?<!-- Close:block:publicationBlock  -->(.*?)</span>.*?<!-- Close:block', re.S)

i = 1
for row in ws.iter_rows():   #对每一行第一列的公司名进行检索以及存储得到的结果
    if 1 == i:
        i = 0
        continue
    company_name = row[0].value
    ticker = row[1].value
    #模拟填写检索信息
    try:
        time.sleep(1)
        elem = driver.find_element_by_id('searchTerm')
        if elem.is_displayed():
            elem.clear()
            elem.send_keys('ti(' + company_name + ') AND ti((billion OR million)) AND ti(invest)')
        else:
            time.sleep(1)
            elem.clear()
            elem.send_keys('ti(' + company_name + ') AND ti((billion OR million)) AND ti(invest)')
        elem.send_keys(Keys.RETURN)
        #等待搜索页面加载10秒
        time.sleep(7)
        
        content = driver.page_source
        items = re.findall(pattern ,content)
        print company_name, '---', len(items)
        for item in items:
            wsres.append([company_name, item[0], item[1]])
        wbres.save('result.xlsx')
    except Exception, e:
        error_log.write(company_name + ': ' + repr(e) + '\n')
        continue
    """
    except Exception:
        #可能某个元素没找到， 重试
        driver.get('http://search.proquest.com.ezproxy.bowdoin.edu/advanced:reset')
        
        driver.find_element_by_id('addRowLink').click() #添加行
        time.sleep(2)
        elem = driver.find_element_by_id('queryTermField')
        if elem.is_displayed():
            elem.send_keys(company_name)  #输入IBM
        else:
            time.sleep(1)
            elem.send_keys(company_name)
        elem = driver.find_element_by_id('queryTermField_0')
        if elem.is_displayed():
            elem.send_keys('billion')  #输入billion
        else:
            time.sleep(1)
            elem.send_keys('billion')
        elem = driver.find_element_by_id('queryTermFieldRight_0')
        if elem.is_displayed():
            elem.send_keys('million') #输入million
        else:
            time.sleep(1)
            elem.send_keys('million')
            
        select = Select(driver.find_element_by_id('fieldsSelect'))
        select.select_by_value('ti')  #选择Document Title
        select = Select(driver.find_element_by_id('fieldsSelect_0'))
        select.select_by_value('ti')    
        
        elem = driver.find_element_by_id('queryTermField_1')
        if elem.is_displayed():
            elem.send_keys('invest')
        else:
            time.sleep(1)
            elem.send_keys('invest')
        select = Select(driver.find_element_by_id('fieldsSelect_1'))
        select.select_by_value('ti')    
        
        
        select = Select(driver.find_element_by_id('select_multiDateRange'))
        select.select_by_value('RANGE')  #选择出版日期为区间
        
        select = Select(driver.find_element_by_id('month2'))  
        select.select_by_value('JANUARY') #开始月份为一月
        select = Select(driver.find_element_by_id('day2'))   
        select.select_by_value('1') #开始日期为1号
        driver.find_element_by_id('year2').send_keys('1996') #1996年起
        
        select = Select(driver.find_element_by_id('month2_0')) 
        select.select_by_value('DECEMBER') #结束月份为12月
        select = Select(driver.find_element_by_id('day2_0'))
        select.select_by_value('31') #结束日期为31日
        driver.find_element_by_id('year2_0').send_keys('2001') #至2001
        
        #等待SourceType选择框加载10秒
        time.sleep(8)
        driver.find_element_by_id('SourceType_Newspapers').click()
        driver.find_element_by_id('SourceType_Wire_Feeds').click()
        
        select = Select(driver.find_element_by_id('itemsPerPage'))
        select.select_by_value('100')
        driver.find_element_by_id('queryTermField').send_keys(Keys.RETURN)
        #等待搜索页面加载10秒
        time.sleep(10)
        
        content = driver.page_source
        items = re.findall(pattern ,content)
        print company_name, '---', len(items)
        for item in items:
            wsres.append([company_name, item[0], item[1]])
        wbres.save('result.xlsx')"""
wbres.save('result.xlsx')
error_log.close()